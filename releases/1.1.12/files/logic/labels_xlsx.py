"""Генерация одного Excel-файла с этикетками (QR, размер, №, наименование) — то же содержимое, что в PDF-этикетке."""
import os
import io
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
from logic.qr_utils import make_remnant_qr_image

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Font = None

# Размер одной этикетки в мм (как в demo.XLS: 100 x 50 мм)
LABEL_WIDTH_MM = 100
LABEL_HEIGHT_MM = 50

# Высота строки в пунктах: 50 мм → pt (1 pt = 1/72 inch)
LABEL_ROW_HEIGHT = round(LABEL_HEIGHT_MM * 72 / 25.4)  # ~142 pt
# Ширина колонок под ~100 мм: в Excel ~1 ед. ≈ 7 px при 96 dpi → сумма ~54 даёт ~100 mm
_COL_WIDTHS = (6, 22, 8, 10, 8)  # №, Наименование, Размер, Текст, QR
# Размер QR в ячейке (пиксели) — вписать в высоту 50 мм: ~50mm ≈ 189 px
QR_CELL_SIZE_PX = min(round(LABEL_HEIGHT_MM * 96 / 25.4), 120)  # ~189, ограничим 120 для читаемости


def _qr_to_png_bytes(unique_number, size_px=320):
    """Возвращает PNG-байты QR-кода остатка (как в PDF)."""
    img = make_remnant_qr_image(unique_number, size_px=size_px)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.read()


def generate_labels_xlsx(remnants, filepath):
    """
    Создаёт один .xlsx с листом «Этикетки»: по одной строке на каждый остаток.
    Колонки: №, Наименование, Размер (мм), подпись «Сканируйте QR», колонка с изображением QR.
    remnants: список dict с ключами unique_number, name, height_mm, width_mm, label_number (опц.).
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Для экспорта этикеток в Excel установите openpyxl: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Этикетки"
    # Заголовки
    ws.append(["№", "Наименование", "Размер (мм)", "Сканируйте QR для истории", ""])
    header_font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = header_font
    # Ширина колонок под этикетку 100 мм (как в demo.XLS)
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    row = 2
    for rem in remnants:
        label_no = rem.get('label_number')
        if label_no is None:
            label_no = rem.get('unique_number')
        name = (rem.get('name') or "")[:80]
        h_mm = rem.get('height_mm') or 0
        w_mm = rem.get('width_mm') or 0
        size_str = "%d x %d" % (h_mm, w_mm)
        ws.cell(row=row, column=1, value=label_no)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=size_str)
        ws.cell(row=row, column=4, value="Сканируйте QR для истории")
        ws.row_dimensions[row].height = LABEL_ROW_HEIGHT
        # Вставка QR в колонку E
        try:
            png_bytes = _qr_to_png_bytes(rem['unique_number'])
            img = XLImage(io.BytesIO(png_bytes))
            img.width = QR_CELL_SIZE_PX
            img.height = QR_CELL_SIZE_PX
            cell_ref = "E%d" % row
            img.anchor = cell_ref
            ws.add_image(img)
        except Exception:
            pass
        row += 1
    wb.save(filepath)


def get_label_export_dir():
    """Папка для сохранения файла этикеток: из настроек пользователя (user_settings), иначе config, иначе папка запуска."""
    try:
        from user_settings import get_labels_dir
        path = get_labels_dir()
        if path:
            return path
    except Exception:
        pass
    path = (getattr(config, 'LABEL_EXPORT_DIR', None) or '').strip()
    if path and os.path.isdir(path):
        return path
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
