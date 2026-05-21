# -*- coding: utf-8 -*-
"""БД MAIN_PROJECT: пользователи (main_users), начальный админ, хелперы статусов заказов.

Пароли в main_users хранятся открытым текстом в колонке password_hash (историческое имя).
"""
import sys
import os
import hashlib
import re
import math

_mp_dir = os.path.dirname(os.path.abspath(__file__))
_root = os.path.dirname(_mp_dir)
if _root not in sys.path:
    sys.path.insert(0, _root)

from db.connection import get_connection, check_tables_exist


def _facade_price_rub_ceil(price):
    """Цены фасадной номенклатуры (профили ₽/м, петли и пр. ₽/шт): целые рубли, округление вверх."""
    if price is None:
        return None
    return int(math.ceil(float(price)))


def _facades_apply_ceil_price_rows(rows, field_name):
    """При чтении из БД: показывать и считать по целым ₽ с округлением вверх (NUMERIC может быть с копейками)."""
    if not rows:
        return rows
    out = []
    for r in rows:
        if r is None:
            continue
        try:
            d = dict(r)
        except Exception:
            d = r
        v = d.get(field_name) if isinstance(d, dict) else None
        if v is not None and isinstance(d, dict):
            try:
                d[field_name] = _facade_price_rub_ceil(v)
            except (TypeError, ValueError):
                pass
        out.append(d)
    return out


def _facades_apply_ceil_price_row(row, field_name):
    if not row:
        return row
    try:
        d = dict(row)
    except Exception:
        return row
    v = d.get(field_name)
    if v is not None:
        try:
            d[field_name] = _facade_price_rub_ceil(v)
        except (TypeError, ValueError):
            pass
    return d


# Роли
ROLE_ADMIN = 'admin'
ROLE_MANAGING = 'managing'
ROLE_MANAGER = 'manager'

# Защищённый администратор: блокировка запрещена (см. set_blocked и UI).
BOSS_PROTECTED_LOGIN = "boss"


def is_boss_protected_login(login) -> bool:
    return str(login or "").strip().lower() == BOSS_PROTECTED_LOGIN

# Подписи как на WEB_SERVICE (admin_users / кабинет): полный доступ, монтаж, офис/производство.
ROLE_LABELS = {
    ROLE_ADMIN: 'Менеджер (полный доступ)',
    ROLE_MANAGING: 'Монтажник',
    ROLE_MANAGER: 'Менеджер',
}


def role_label_desktop(role):
    """Подпись роли в офисной программе: администратор — «Админ», остальные как в ROLE_LABELS."""
    r = str(role or "").strip()
    if r == ROLE_ADMIN:
        return "Админ"
    return ROLE_LABELS.get(r, r or "—")


# Откуда создана учётная запись: офисная программа или регистрация на сайте (WEB_SERVICE).
ACCOUNT_ORIGIN_DESKTOP = 'desktop'
ACCOUNT_ORIGIN_WEB = 'web'


def account_origin_ru(origin) -> str:
    """Краткая подпись источника учётной записи для таблиц в программе."""
    o = (origin or ACCOUNT_ORIGIN_DESKTOP).strip().lower()
    if o == ACCOUNT_ORIGIN_WEB:
        return 'WEB_SERVICE (сайт)'
    return 'Программа (офис)'

# Статусы заказа (расширенные)
ORDER_STATUS_DRAFT = 'draft'
ORDER_STATUS_PAID = 'paid'
ORDER_STATUS_IN_PROGRESS = 'in_progress'
ORDER_STATUS_MADE = 'made'
ORDER_STATUS_CHECKED_QR = 'checked_qr'
ORDER_STATUS_SHIPPED = 'shipped'

ORDER_STATUS_RU = {
    ORDER_STATUS_DRAFT: 'Просчет',
    ORDER_STATUS_PAID: 'Оплачен',
    ORDER_STATUS_IN_PROGRESS: 'В работе',
    ORDER_STATUS_MADE: 'Изготовлен',
    ORDER_STATUS_CHECKED_QR: 'Проверен по QR',
    ORDER_STATUS_SHIPPED: 'Отгружен',
    'completed': 'Выполнен',  # обратная совместимость
    'cancelled': 'Отменён',
}


def order_status_to_ru(status):
    s = (str(status).strip() if status is not None else "")
    if not s:
        return "—"
    return str(ORDER_STATUS_RU.get(s.lower(), s))


# Печать карт раскроя (PDF), этикетки и пр. — с «Оплачен» и по цепочке дальше (в т.ч. «В работе»).
ORDER_STATUS_ALLOWS_PRODUCTION_PRINT = frozenset(
    {
        ORDER_STATUS_PAID,
        ORDER_STATUS_IN_PROGRESS,
        ORDER_STATUS_MADE,
        ORDER_STATUS_CHECKED_QR,
        ORDER_STATUS_SHIPPED,
        "completed",
    }
)


def order_status_allows_production_print(status) -> bool:
    return str(status or "").strip() in ORDER_STATUS_ALLOWS_PRODUCTION_PRINT


# Правка расчёта / состава заказа — только до передачи в производство (не «В работе» и не позже).
ORDER_STATUS_ALLOWS_BUNDLE_EDIT = frozenset({ORDER_STATUS_DRAFT, ORDER_STATUS_PAID})


def order_status_allows_bundle_edit(status) -> bool:
    return str(status or "").strip() in ORDER_STATUS_ALLOWS_BUNDLE_EDIT


def _ensure_main_app_meta_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS main_app_meta (
            k TEXT PRIMARY KEY,
            v TEXT NOT NULL DEFAULT ''
        )
    """)


def _main_meta_get(cur, key):
    cur.execute("SELECT v FROM main_app_meta WHERE k = %s", (key,))
    row = cur.fetchone()
    return row["v"] if row else None


def _main_meta_set(cur, key, value):
    cur.execute(
        """
        INSERT INTO main_app_meta (k, v) VALUES (%s, %s)
        ON CONFLICT (k) DO UPDATE SET v = EXCLUDED.v
        """,
        (key, value),
    )


def _ensure_main_users_account_origin_column(cur):
    """Колонка account_origin: desktop — регистрация из главной программы, web — с WEB_SERVICE."""
    cur.execute(
        """
        ALTER TABLE main_users
        ADD COLUMN IF NOT EXISTS account_origin VARCHAR(32) NOT NULL DEFAULT 'desktop'
        """
    )


def _migrate_main_users_password_column_to_text(cur):
    cur.execute("""
        SELECT data_type, character_maximum_length
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = 'main_users' AND column_name = 'password_hash'
    """)
    row = cur.fetchone()
    if not row:
        return
    if (row.get("data_type") or "").lower() == "text":
        return
    cur.execute("ALTER TABLE main_users ALTER COLUMN password_hash TYPE TEXT")


def _one_time_reset_users_plain_password(cur, admin_login, admin_password_plain):
    """
    Один раз на БД: удалить всех, кроме админа из cfg, выставить пароль админа открытым текстом.
    Возвращает True, если миграция только что выполнена (нужно подчистить Django auth_user).
    """
    _ensure_main_app_meta_table(cur)
    if _main_meta_get(cur, "main_users_plain_v1") == "1":
        return False
    al = (admin_login or "").strip()
    ap = admin_password_plain or ""
    cur.execute(
        "DELETE FROM main_users WHERE LOWER(TRIM(login)) <> LOWER(TRIM(%s))",
        (al,),
    )
    cur.execute(
        """
        SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s)) ORDER BY id
        """,
        (al,),
    )
    dup_rows = cur.fetchall() or []
    if len(dup_rows) > 1:
        keep0 = int(dup_rows[0]["id"])
        for r in dup_rows[1:]:
            cur.execute("DELETE FROM main_users WHERE id = %s", (int(r["id"]),))
        dup_rows = dup_rows[:1]
    row = dup_rows[0] if dup_rows else None
    if row:
        cur.execute(
            """
            UPDATE main_users
            SET password_hash = %s, role = %s, approved = TRUE, blocked = FALSE,
                login = %s,
                account_origin = %s,
                surname = COALESCE(NULLIF(TRIM(surname), ''), 'Admin'),
                name = COALESCE(NULLIF(TRIM(name), ''), 'Administrator')
            WHERE id = %s
            """,
            (ap, ROLE_ADMIN, al, ACCOUNT_ORIGIN_DESKTOP, int(row["id"])),
        )
    else:
        cur.execute(
            """INSERT INTO main_users (surname, name, login, password_hash, role, approved, blocked, account_origin)
               VALUES (%s, %s, %s, %s, %s, TRUE, FALSE, %s)""",
            ("Admin", "Administrator", al, ap, ROLE_ADMIN, ACCOUNT_ORIGIN_DESKTOP),
        )
    _main_meta_set(cur, "main_users_plain_v1", "1")
    return True


def _ensure_boss_admin_account(cur):
    """Гарантировать учётку boss / dima (если ещё нет в main_users)."""
    cur.execute(
        "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s))",
        (BOSS_PROTECTED_LOGIN,),
    )
    if cur.fetchone():
        return
    cur.execute(
        """INSERT INTO main_users (surname, name, login, password_hash, role, approved, blocked, account_origin)
           VALUES (%s, %s, %s, %s, %s, TRUE, FALSE, %s)""",
        ("Босс", "Дима", BOSS_PROTECTED_LOGIN, "dima", ROLE_ADMIN, ACCOUNT_ORIGIN_DESKTOP),
    )


def _prune_django_users_keep_login(keep_login: str) -> None:
    """Удалить лишних пользователей Django (после сброса main_users)."""
    lg = (keep_login or "").strip()
    if not lg:
        return
    try:
        mp = os.path.join(_root, "montazh_portal")
        if mp not in sys.path:
            sys.path.insert(0, mp)
        os.environ.setdefault("DJANGO_SETTINGS_MODULE", "montazh_portal.settings")
        import django  # noqa: WPS433
        from django.conf import settings as django_settings  # noqa: WPS433

        if not django_settings.configured:
            django.setup()
        from django.contrib.auth.models import User  # noqa: WPS433

        User.objects.exclude(username__iexact=lg).delete()
    except Exception:
        pass


def ensure_main_tables():
    """Создать main_users если нет, заполнить начального админа из cfg."""
    from cfg_loader import app_cfg, get_cfg_string
    cfg = app_cfg()
    admin_login = get_cfg_string(cfg, 'admin', 'login', 'admin')
    admin_password = get_cfg_string(cfg, 'admin', 'password', 'admin')

    existing = check_tables_exist(['main_users'])
    did_plain_migrate = False
    with get_connection() as conn:
        with conn.cursor() as cur:
            if 'main_users' not in existing:
                cur.execute("""
                    CREATE TABLE main_users (
                        id SERIAL PRIMARY KEY,
                        surname VARCHAR(255) NOT NULL DEFAULT '',
                        name VARCHAR(255) NOT NULL DEFAULT '',
                        login VARCHAR(128) UNIQUE NOT NULL,
                        password_hash TEXT NOT NULL,
                        role VARCHAR(32) NOT NULL DEFAULT 'manager',
                        approved BOOLEAN NOT NULL DEFAULT FALSE,
                        blocked BOOLEAN NOT NULL DEFAULT FALSE,
                        account_origin VARCHAR(32) NOT NULL DEFAULT 'desktop',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            _migrate_main_users_password_column_to_text(cur)
            _ensure_main_users_account_origin_column(cur)
            did_plain_migrate = bool(
                _one_time_reset_users_plain_password(cur, admin_login, admin_password)
            )
            cur.execute(
                "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s))",
                (admin_login,),
            )
            if not cur.fetchone():
                cur.execute(
                    """INSERT INTO main_users (surname, name, login, password_hash, role, approved, blocked, account_origin)
                       VALUES (%s, %s, %s, %s, %s, TRUE, FALSE, %s)""",
                    ('Admin', 'Administrator', admin_login, admin_password, ROLE_ADMIN, ACCOUNT_ORIGIN_DESKTOP)
                )
            _ensure_boss_admin_account(cur)
    if did_plain_migrate:
        _prune_django_users_keep_login(admin_login)


def get_user_by_login(login):
    q = (login or "").strip()
    if not q:
        return None
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, surname, name, login, password_hash, role, approved, blocked, account_origin, created_at
                   FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s))""",
                (q,),
            )
            return cur.fetchone()


def check_password(user, password):
    if not user:
        return False
    stored = (user.get('password_hash') or '')
    return stored == (password if password is None else str(password))


def create_user(surname, name, login, password, role=ROLE_MANAGER, account_origin=ACCOUNT_ORIGIN_DESKTOP):
    """Регистрация: approved=False. Возвращает id или None при ошибке (логин занят)."""
    ao = (account_origin or ACCOUNT_ORIGIN_DESKTOP).strip().lower()
    if ao not in (ACCOUNT_ORIGIN_DESKTOP, ACCOUNT_ORIGIN_WEB):
        ao = ACCOUNT_ORIGIN_DESKTOP
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s))",
                ((login or "").strip(),),
            )
            if cur.fetchone():
                return None
            cur.execute(
                """INSERT INTO main_users (surname, name, login, password_hash, role, approved, blocked, account_origin)
                   VALUES (%s, %s, %s, %s, %s, FALSE, FALSE, %s) RETURNING id""",
                (
                    (surname or '').strip(),
                    (name or '').strip(),
                    (login or '').strip(),
                    (password or ''),
                    (role or ROLE_MANAGER),
                    ao,
                ),
            )
            row = cur.fetchone()
            return row['id'] if row else None


def create_admin_user(surname, name, login, password, account_origin=ACCOUNT_ORIGIN_DESKTOP):
    """Создать администратора из админки: сразу подтверждён. None если логин занят."""
    ao = (account_origin or ACCOUNT_ORIGIN_DESKTOP).strip().lower()
    if ao not in (ACCOUNT_ORIGIN_DESKTOP, ACCOUNT_ORIGIN_WEB):
        ao = ACCOUNT_ORIGIN_DESKTOP
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s))",
                ((login or "").strip(),),
            )
            if cur.fetchone():
                return None
            cur.execute(
                """INSERT INTO main_users (surname, name, login, password_hash, role, approved, blocked, account_origin)
                   VALUES (%s, %s, %s, %s, %s, TRUE, FALSE, %s) RETURNING id""",
                (
                    (surname or "").strip(),
                    (name or "").strip(),
                    (login or "").strip(),
                    (password or ""),
                    ROLE_ADMIN,
                    ao,
                ),
            )
            row = cur.fetchone()
            return row["id"] if row else None


def get_unapproved_count():
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS c FROM main_users WHERE approved = FALSE AND blocked = FALSE")
            return (cur.fetchone() or {}).get('c', 0)


def get_all_users():
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, surname, name, login, password_hash, role, approved, blocked, account_origin, created_at
                   FROM main_users ORDER BY id"""
            )
            return cur.fetchall()


def get_unapproved_users():
    """Пользователи, ожидающие подтверждения (не заблокированы)."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, surname, name, login, role, approved, blocked, account_origin, created_at
                   FROM main_users
                   WHERE approved = FALSE AND blocked = FALSE
                   ORDER BY id"""
            )
            return cur.fetchall()


def get_user_by_id(user_id):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, surname, name, login, password_hash, role, approved, blocked, account_origin, created_at
                   FROM main_users WHERE id = %s""",
                (int(user_id),),
            )
            return cur.fetchone()


def is_boss_protected_user_id(user_id) -> bool:
    u = get_user_by_id(int(user_id))
    return bool(u and is_boss_protected_login(u.get("login")))


def set_user_password_plain(user_id, plain_password):
    """Сохранить пароль в main_users открытым текстом. Возвращает True если строка обновлена."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE main_users SET password_hash = %s WHERE id = %s",
                (plain_password if plain_password is None else str(plain_password), int(user_id)),
            )
            return cur.rowcount > 0


def sync_montazh_django_password_after_main_users_change(login, plain_password):
    """
    После смены пароля в main_users обновить Django auth_user (тот же алгоритм, что auth_backend).
    Иначе пароль в БД main_users новый, а в auth_user старый — путаница при входе в /montazh.
    """
    lg = (login or "").strip()
    if not lg:
        return
    try:
        mp = os.path.join(_root, "montazh_portal")
        if mp not in sys.path:
            sys.path.insert(0, mp)
        os.environ.setdefault("DJANGO_SETTINGS_MODULE", "montazh_portal.settings")
        import django  # noqa: WPS433
        from django.conf import settings as django_settings  # noqa: WPS433

        if not django_settings.configured:
            django.setup()
        from django.contrib.auth.models import User  # noqa: WPS433
        from accounts.password_utils import normalize_password  # noqa: WPS433

        u = User.objects.filter(username__iexact=lg).first()
        if not u:
            return
        u.set_password(normalize_password(plain_password) or (plain_password or ""))
        u.save()
    except Exception:
        pass


def sync_montazh_django_username_change(old_login, new_login):
    """После смены логина в main_users обновить username в Django auth_user, если есть."""
    old_l = (old_login or "").strip()
    new_l = (new_login or "").strip()
    if not old_l or not new_l or old_l.lower() == new_l.lower():
        return
    try:
        mp = os.path.join(_root, "montazh_portal")
        if mp not in sys.path:
            sys.path.insert(0, mp)
        os.environ.setdefault("DJANGO_SETTINGS_MODULE", "montazh_portal.settings")
        import django  # noqa: WPS433
        from django.conf import settings as django_settings  # noqa: WPS433

        if not django_settings.configured:
            django.setup()
        from django.contrib.auth.models import User  # noqa: WPS433

        u = User.objects.filter(username__iexact=old_l).first()
        if not u:
            return
        if User.objects.filter(username__iexact=new_l).exclude(pk=u.pk).exists():
            return
        u.username = new_l
        u.save()
    except Exception:
        pass


def update_user_credentials_admin(user_id, surname, name, login, password_plain=None):
    """
    Админ: сохранить фамилию, имя, логин; пароль — только если password_plain не пустой.
    Возвращает (True, None) или (False, текст_ошибки).
    """
    uid = int(user_id)
    row = get_user_by_id(uid)
    if not row:
        return False, "Пользователь не найден."
    new_login = (login or "").strip()
    if not new_login:
        return False, "Логин не может быть пустым."
    old_login = (row.get("login") or "").strip()
    sn = ((surname or "").strip())[:255]
    nm = ((name or "").strip())[:255]
    pwd = password_plain
    change_pwd = pwd is not None and str(pwd).strip() != ""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s)) AND id <> %s",
                (new_login, uid),
            )
            if cur.fetchone():
                return False, "Пользователь с таким логином уже существует."
            if change_pwd:
                cur.execute(
                    "UPDATE main_users SET surname = %s, name = %s, login = %s, password_hash = %s WHERE id = %s",
                    (sn, nm, new_login, str(pwd), uid),
                )
            else:
                cur.execute(
                    "UPDATE main_users SET surname = %s, name = %s, login = %s WHERE id = %s",
                    (sn, nm, new_login, uid),
                )
            if cur.rowcount < 1:
                return False, "Не удалось сохранить изменения."
    if old_login.lower() != new_login.lower():
        sync_montazh_django_username_change(old_login, new_login)
    if change_pwd:
        sync_montazh_django_password_after_main_users_change(new_login, pwd)
    return True, None


def self_service_update_profile(
    user_id,
    new_login,
    old_password="",
    new_password="",
    new_password_confirm="",
    surname=None,
    name=None,
):
    """
    Личный кабинет: логин; пароль — только если заполнены все три поля (старый, новый, повтор).
    Фамилию и имя может менять только пользователь с ролью admin (остальные — поля игнорируются, в БД остаются старые).
    Возвращает (True, None) или (False, текст_ошибки).
    """
    uid = int(user_id)
    row = get_user_by_id(uid)
    if not row:
        return False, "Пользователь не найден."
    new_login = (new_login or "").strip()
    if not new_login:
        return False, "Введите логин."
    if str(row.get("role") or "").strip() == ROLE_ADMIN:
        sn = ((surname if surname is not None else row.get("surname")) or "").strip()[:255]
        nm = ((name if name is not None else row.get("name")) or "").strip()[:255]
    else:
        sn = ((row.get("surname") or "").strip())[:255]
        nm = ((row.get("name") or "").strip())[:255]
    old_l = (row.get("login") or "").strip()
    op = old_password if old_password is None else str(old_password)
    np = new_password if new_password is None else str(new_password)
    npc = new_password_confirm if new_password_confirm is None else str(new_password_confirm)
    pwd_any = bool(str(op).strip() or str(np).strip() or str(npc).strip())
    if pwd_any:
        if not str(op).strip() or not str(np).strip() or not str(npc).strip():
            return False, "Для смены пароля укажите старый пароль и новый дважды."
        if np != npc:
            return False, "Новые пароли не совпадают."
        if not check_password(row, op):
            return False, "Старый пароль указан неверно."
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM main_users WHERE LOWER(TRIM(login)) = LOWER(TRIM(%s)) AND id <> %s",
                (new_login, uid),
            )
            if cur.fetchone():
                return False, "Пользователь с таким логином уже существует."
            if pwd_any:
                cur.execute(
                    "UPDATE main_users SET surname = %s, name = %s, login = %s, password_hash = %s WHERE id = %s",
                    (sn, nm, new_login, np, uid),
                )
            else:
                cur.execute(
                    "UPDATE main_users SET surname = %s, name = %s, login = %s WHERE id = %s",
                    (sn, nm, new_login, uid),
                )
            if cur.rowcount < 1:
                return False, "Не удалось сохранить изменения."
    if old_l.lower() != new_login.lower():
        sync_montazh_django_username_change(old_l, new_login)
    if pwd_any:
        sync_montazh_django_password_after_main_users_change(new_login, np)
    return True, None


def set_approved(user_id, approved=True):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE main_users SET approved = %s WHERE id = %s", (bool(approved), int(user_id)))


def set_blocked(user_id, blocked=True):
    if blocked and is_boss_protected_user_id(user_id):
        return False
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE main_users SET blocked = %s WHERE id = %s", (bool(blocked), int(user_id)))
            return cur.rowcount > 0


def admin_update_user_identity(user_id, surname=None, name=None, role=None):
    """
    Обновить ФИО и роль пользователя из админ-панели.
    Возвращает True, если запись обновлена.
    """
    fields = []
    params = []
    if surname is not None:
        fields.append("surname = %s")
        params.append(((surname or "").strip())[:255])
    if name is not None:
        fields.append("name = %s")
        params.append(((name or "").strip())[:255])
    if role is not None:
        rv = ((role or "").strip())[:32]
        if not rv:
            pass
        elif rv not in (ROLE_ADMIN, ROLE_MANAGING, ROLE_MANAGER):
            raise ValueError("Недопустимая роль.")
        else:
            fields.append("role = %s")
            params.append(rv)
    if not fields:
        return False
    with get_connection() as conn:
        with conn.cursor() as cur:
            # Имена колонок из белого списка выше; значения только через параметры.
            sql = "UPDATE main_users SET " + ", ".join(fields) + " WHERE id = %s"
            cur.execute(sql, tuple(params + [int(user_id)]))
            return cur.rowcount > 0


# --- Фасады: профили и петли ---

def ensure_facades_tables():
    """Создать таблицы фасадов: профили, петли, винты, площадки, уплотнитель, уголки."""
    existing = check_tables_exist([
        'facades_profile', 'facades_hinges',
        'facades_screws', 'facades_plates', 'facades_seal', 'facades_corners',
        'facades_angle_seal',
    ])
    with get_connection() as conn:
        with conn.cursor() as cur:
            if 'facades_profile' not in existing:
                cur.execute("""
                    CREATE TABLE facades_profile (
                        id SERIAL PRIMARY KEY,
                        series VARCHAR(255) DEFAULT '',
                        name VARCHAR(255) NOT NULL DEFAULT '',
                        color VARCHAR(255) DEFAULT '',
                        supplier VARCHAR(255) DEFAULT '',
                        price_per_meter NUMERIC(12,2) NOT NULL DEFAULT 0,
                        photo_number VARCHAR(64) DEFAULT '',
                        link TEXT DEFAULT '',
                        handle VARCHAR(255) DEFAULT '',
                        price_updated_at TIMESTAMP DEFAULT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            # Добавить колонку link в существующую таблицу профилей, если её нет
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'facades_profile' AND column_name = 'link'
            """)
            if not cur.fetchone():
                cur.execute("ALTER TABLE facades_profile ADD COLUMN link TEXT DEFAULT ''")
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'facades_profile' AND column_name = 'price_updated_at'
            """)
            if not cur.fetchone():
                cur.execute(
                    "ALTER TABLE facades_profile ADD COLUMN price_updated_at TIMESTAMP DEFAULT NULL"
                )
            if 'facades_hinges' not in existing:
                cur.execute("""
                    CREATE TABLE facades_hinges (
                        id SERIAL PRIMARY KEY,
                        number VARCHAR(32) DEFAULT '',
                        series VARCHAR(255) DEFAULT '',
                        name VARCHAR(255) NOT NULL DEFAULT '',
                        color VARCHAR(255) DEFAULT '',
                        supplier VARCHAR(255) DEFAULT '',
                        price NUMERIC(12,2) DEFAULT NULL,
                        price_updated_at TIMESTAMP DEFAULT NULL,
                        item VARCHAR(255) DEFAULT '',
                        photo_number VARCHAR(64) DEFAULT '',
                        link TEXT DEFAULT '',
                        catalog_primary BOOLEAN DEFAULT FALSE,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            # Миграции: добавить недостающие колонки в facades_hinges
            for col_name, col_def in [
                ('number', "VARCHAR(32) DEFAULT ''"),
                ('photo_number', "VARCHAR(64) DEFAULT ''"),
                ('series', "VARCHAR(255) DEFAULT ''"),
                ('color', "VARCHAR(255) DEFAULT ''"),
                ('supplier', "VARCHAR(255) DEFAULT ''"),
                ('item', "VARCHAR(255) DEFAULT ''"),
                ('catalog_primary', "BOOLEAN DEFAULT FALSE"),
            ]:
                cur.execute("""
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'facades_hinges' AND column_name = %s
                """, (col_name,))
                if not cur.fetchone():
                    cur.execute("ALTER TABLE facades_hinges ADD COLUMN " + col_name + " " + col_def)
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'facades_hinges' AND column_name = 'catalog_primary'
            """)
            if cur.fetchone():
                cur.execute(
                    "UPDATE facades_hinges SET catalog_primary = FALSE "
                    "WHERE catalog_primary IS NULL"
                )
            # Таблицы: винты, площадки, уплотнитель, уголки (одинаковая структура)
            _facades_item_schema = """
                id SERIAL PRIMARY KEY,
                number VARCHAR(32) DEFAULT '',
                series VARCHAR(255) DEFAULT '',
                name VARCHAR(255) NOT NULL DEFAULT '',
                color VARCHAR(255) DEFAULT '',
                supplier VARCHAR(255) DEFAULT '',
                price NUMERIC(12,2) DEFAULT NULL,
                price_updated_at TIMESTAMP DEFAULT NULL,
                photo_number VARCHAR(64) DEFAULT '',
                link TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            """
            for tbl in ('facades_screws', 'facades_plates', 'facades_seal', 'facades_corners'):
                if tbl not in existing:
                    cur.execute("CREATE TABLE " + tbl + " (" + _facades_item_schema + ")")

            # Сводная таблица "уголки_уплотнитель" для админа:
            # - Угловой соединитель: продаётся по штукам
            # - Уплотнитель: продаётся по метрам
            if 'facades_angle_seal' not in existing:
                cur.execute("""
                    CREATE TABLE facades_angle_seal (
                        id SERIAL PRIMARY KEY,
                        item_type VARCHAR(128) NOT NULL DEFAULT '',
                        variant VARCHAR(128) NOT NULL DEFAULT '',
                        unit VARCHAR(16) NOT NULL DEFAULT '',
                        price NUMERIC(12,2) DEFAULT NULL,
                        price_updated_at TIMESTAMP DEFAULT NULL,
                        link TEXT DEFAULT '',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)

            cur.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_facades_angle_seal_item
                ON facades_angle_seal (item_type, variant)
                """
            )

            # Seed: то, что вы просили добавить в таблицу админа
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Угловой соединитель', 'F3-021', 'шт', 80, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Угловой соединитель' AND variant = 'F3-021'
                )
                """
            )
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Угловой соединитель', 'F3-031', 'шт', 100, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Угловой соединитель' AND variant = 'F3-031'
                )
                """
            )
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Уплотнитель', 'черный', 'м', 45, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Уплотнитель' AND variant = 'черный'
                )
                """
            )
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Уплотнитель', 'прозрачный', 'м', 45, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Уплотнитель' AND variant = 'прозрачный'
                )
                """
            )
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Винт', 'серебро', 'шт', 2, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Винт' AND variant = 'серебро'
                )
                """
            )
            cur.execute(
                """
                INSERT INTO facades_angle_seal (item_type, variant, unit, price, link)
                SELECT 'Винт', 'золото', 'шт', 2, ''
                WHERE NOT EXISTS (
                    SELECT 1 FROM facades_angle_seal
                    WHERE item_type = 'Винт' AND variant = 'золото'
                )
                """
            )
            conn.commit()


def facades_get_all_profiles(series=None, name=None, color=None, supplier=None):
    """Список профилей с опциональными фильтрами."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            q = (
                "SELECT id, series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at "
                "FROM facades_profile WHERE 1=1"
            )
            params = []
            if series and str(series).strip():
                params.append('%' + str(series).strip() + '%')
                q += " AND series ILIKE %s"
            if name and str(name).strip():
                params.append('%' + str(name).strip() + '%')
                q += " AND name ILIKE %s"
            if color and str(color).strip():
                params.append('%' + str(color).strip() + '%')
                q += " AND color ILIKE %s"
            if supplier and str(supplier).strip():
                params.append('%' + str(supplier).strip() + '%')
                q += " AND supplier ILIKE %s"
            q += " ORDER BY series, name, color"
            cur.execute(q, params or None)
            return _facades_apply_ceil_price_rows(cur.fetchall(), "price_per_meter")


def facades_update_profile_price(profile_id, price_per_meter):
    pm = _facade_price_rub_ceil(price_per_meter)
    if pm is None:
        return
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE facades_profile SET price_per_meter = %s, price_updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                (pm, int(profile_id)),
            )


def facades_add_profile_variant_from_base(base_profile_id, color, price_per_meter):
    """Создать вариант профиля по цвету на основе существующего профиля."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT series, name, supplier, photo_number, link, handle
                   FROM facades_profile WHERE id = %s""",
                (int(base_profile_id),)
            )
            base = cur.fetchone()
            if not base:
                return None
            cur.execute(
                """INSERT INTO facades_profile (series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP) RETURNING id""",
                (
                    str(base.get('series') or '').strip(),
                    str(base.get('name') or '').strip(),
                    str(color or '').strip(),
                    str(base.get('supplier') or '').strip(),
                    (_facade_price_rub_ceil(price_per_meter) if price_per_meter is not None else 0) or 0,
                    str(base.get('photo_number') or '').strip(),
                    str(base.get('link') or '').strip(),
                    str(base.get('handle') or '').strip(),
                )
            )
            row = cur.fetchone()
            return row['id'] if row else None


def _norm_article(s):
    return re.sub(r'[^a-z0-9а-я]+', '', str(s or '').lower())


def _split_color_tokens(color_text):
    text = str(color_text or '').strip()
    if not text:
        return []
    parts = [p.strip() for p in re.split(r'[,;/]+', text) if str(p).strip()]
    # уникальные в исходном порядке
    out = []
    seen = set()
    for p in parts:
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def _photo_numeric_base(photo_val):
    """Лидирующие цифры из ячейки «№ фото» (например 15 → файлы 15_champagne.png)."""
    s = str(photo_val or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d+)", s)
    return m.group(1) if m else ""


_PROFILE_COLOR_SLUG_RU = {
    "шампань": "champagne",
    "золото": "gold",
    "коньяк": "cognac",
    "серебро": "silver",
    "белый": "white",
    "черный": "black",
    "чёрный": "black",
    "бронза": "bronze",
    "кофе": "coffee",
    "коричневый": "brown",
    "графит": "graphite",
    "антрацит": "anthracite",
    "хром": "chrome",
    "сатин": "satin",
    "никель": "nickel",
    "матовый": "matte",
    "матов": "matte",
}


def _profile_color_slug(token):
    """Латинский суффикс для имени файла FASAD/img/{base}_{slug}.png."""
    t = (token or "").strip().lower()
    if not t:
        return "x"
    if t in _PROFILE_COLOR_SLUG_RU:
        return _PROFILE_COLOR_SLUG_RU[t]
    for ru, en in _PROFILE_COLOR_SLUG_RU.items():
        if t.startswith(ru):
            return en
    asc = re.sub(r"[^a-z0-9]+", "_", t)
    asc = asc.strip("_")
    if asc:
        return asc[:48]
    return "c_" + hashlib.md5(t.encode("utf-8")).hexdigest()[:8]


_PXL_COL = {
    "series": 1,
    "name": 2,
    "color": 3,
    "supplier": 4,
    "price": 5,
    "photo": 6,
    "link": 7,
}


def _iter_profil_excel_profile_rows(file_path):
    """
    Лист «Профиль» (Profil_new.xlsx): фиксированные колонки A–H.
    Якорная строка (серия+название) + продолжения с пустыми серия/название.
    Несколько цветов в одной ячейке → несколько позиций, photo_number = «{№фото}_{slug}».
    Возвращает (список dict, error_message или None).
    """
    try:
        import openpyxl
    except ImportError:
        return [], "Установите openpyxl: pip install openpyxl"
    if not os.path.isfile(file_path):
        return [], "Файл не найден: %s" % file_path
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = None
            for sname in wb.sheetnames:
                if sname and "профиль" in sname.lower():
                    ws = wb[sname]
                    break
            if ws is None:
                ws = wb.active
            rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0, values_only=True))
        finally:
            wb.close()
    except Exception as e:
        return [], str(e)
    if len(rows) < 2:
        return [], "Лист «Профиль» пустой"

    anchor_series = ""
    anchor_name = ""
    anchor_link = ""
    out = []

    for row in rows[1:]:
        if not row:
            continue
        row = list(row)
        while len(row) < 8:
            row.append(None)

        def _cell(key):
            i = _PXL_COL[key]
            return row[i] if i < len(row) else None

        s_series = str(_cell("series") or "").strip()
        s_name = str(_cell("name") or "").strip()
        link_cell = _cell("link")
        link_str = str(link_cell or "").strip() if link_cell else ""

        if s_series and s_name:
            anchor_series, anchor_name = s_series, s_name
            if link_str:
                anchor_link = link_str
        elif not (anchor_series and anchor_name):
            continue

        use_link = link_str or anchor_link
        color_combined = str(_cell("color") or "").strip()
        if not color_combined:
            continue

        try:
            price_val = float(_cell("price")) if _cell("price") is not None else 0.0
        except (TypeError, ValueError):
            price_val = 0.0

        photo_base = _photo_numeric_base(_cell("photo"))
        if not photo_base:
            continue

        supplier = str(_cell("supplier") or "").strip()
        tokens = _split_color_tokens(color_combined)
        if not tokens:
            tokens = [color_combined]

        for tok in tokens:
            slug = _profile_color_slug(tok)
            photo_db = "%s_%s" % (photo_base, slug)
            out.append(
                {
                    "series": anchor_series,
                    "name": anchor_name,
                    "color": tok.strip(),
                    "supplier": supplier,
                    "price_per_meter": _facade_price_rub_ceil(price_val) or 0,
                    "photo_number": photo_db,
                    "link": use_link,
                    "handle": "0",
                }
            )

    if not out:
        return [], "Не удалось разобрать строки профилей (проверьте лист «Профиль» и колонки A–H)."
    return out, None


def _norm_excel_header_cell(v):
    s = str(v or "").strip().lower().replace("\n", " ")
    return re.sub(r"\s+", " ", s)


def _iter_simple_profile_price_rows(file_path):
    """
    Простой шаблон: первая строка — заголовки с колонками Серия, Название, Цвет, Цена/м (или похожими).
    Данные со 2-й строки. Совпадение колонок по подстрокам в заголовке.
    Возвращает (список dict с ключами series, name, color, price_per_meter, error или None).
    """
    try:
        import openpyxl
    except ImportError:
        return [], "Установите openpyxl: pip install openpyxl"
    if not os.path.isfile(file_path):
        return [], "Файл не найден: %s" % file_path
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0, values_only=True))
        finally:
            wb.close()
    except Exception as e:
        return [], str(e)
    if len(rows) < 2:
        return [], "В файле нет строк данных (нужны заголовок и хотя бы одна строка)"

    hdr = [_norm_excel_header_cell(c) for c in (rows[0] or ())]

    def col_idx(*substrings):
        """Первая колонка, в заголовке которой есть все подстроки."""
        for i, h in enumerate(hdr):
            if not h:
                continue
            if all(sub in h for sub in substrings):
                return i
        return None

    i_ser = col_idx("серия")
    i_name = col_idx("название")
    if i_name is None:
        i_name = col_idx("наименование")
    i_col = col_idx("цвет")
    i_price = col_idx("цена", "/м")
    if i_price is None:
        i_price = col_idx("цена", "м")
    if i_price is None:
        for i, h in enumerate(hdr):
            if h in ("цена", "цена/м", "цена за м", "руб/м") or (
                "цена" in h and ("м" in h or "/" in h)
            ):
                i_price = i
                break

    if i_ser is None or i_name is None or i_col is None or i_price is None:
        return [], (
            "Нужны колонки в первой строке: Серия, Название, Цвет, Цена/м "
            "(или формат Profil_new — лист «Профиль»)"
        )

    out = []
    for row in rows[1:]:
        if not row:
            continue
        row = list(row)
        while len(row) <= max(i_ser, i_name, i_col, i_price):
            row.append(None)

        def get(ci):
            return row[ci] if ci < len(row) else None

        series = str(get(i_ser) or "").strip()
        name = str(get(i_name) or "").strip()
        color = str(get(i_col) or "").strip()
        if not series or not name or not color:
            continue
        raw_p = get(i_price)
        try:
            price_val = (
                _facade_price_rub_ceil(raw_p)
                if raw_p is not None and str(raw_p).strip() != ""
                else None
            )
        except (TypeError, ValueError):
            price_val = None
        if price_val is None:
            continue
        out.append(
            {
                "series": series,
                "name": name,
                "color": color,
                "price_per_meter": price_val,
            }
        )

    if not out:
        return [], "Нет ни одной строки с заполненными серия, название, цвет и числовой ценой"
    return out, None


def _facades_apply_profile_price_updates(rows_flat):
    """Обновить price_per_meter и price_updated_at по списку dict (series, name, color, price_per_meter)."""
    updated = 0
    not_found = []
    with get_connection() as conn:
        with conn.cursor() as cur:
            for r in rows_flat:
                try:
                    pm = _facade_price_rub_ceil(r["price_per_meter"])
                except (TypeError, ValueError):
                    continue
                if pm is None:
                    continue
                cur.execute(
                    """UPDATE facades_profile SET price_per_meter = %s, price_updated_at = CURRENT_TIMESTAMP
                       WHERE lower(trim(series)) = lower(trim(%s))
                         AND lower(trim(name)) = lower(trim(%s))
                         AND lower(trim(color)) = lower(trim(%s))""",
                    (
                        pm,
                        r["series"],
                        r["name"],
                        r["color"],
                    ),
                )
                if cur.rowcount == 0:
                    not_found.append("%s | %s | %s" % (r["series"], r["name"], r["color"]))
                else:
                    updated += cur.rowcount
    return {
        "error": None,
        "updated": updated,
        "not_found": not_found,
        "source_rows": len(rows_flat),
    }


def facades_write_profiles_demo_excel(file_path):
    """
    Экспорт текущих профилей в простой .xlsx: Серия, Название, Цвет, Цена/м.
    Пользователь меняет цены и импортирует файл обратно.
    """
    try:
        import openpyxl
    except ImportError:
        return False, "Установите openpyxl: pip install openpyxl"
    profiles = facades_get_all_profiles()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Профили"
        ws.append(["Серия", "Название", "Цвет", "Цена/м"])
        for p in profiles:
            try:
                price = _facade_price_rub_ceil(p.get("price_per_meter") or 0) or 0
            except (TypeError, ValueError):
                price = 0
            ws.append(
                [
                    str(p.get("series") or ""),
                    str(p.get("name") or ""),
                    str(p.get("color") or ""),
                    price,
                ]
            )
        wb.save(file_path)
        return True, None
    except Exception as e:
        return False, str(e)


def facades_expand_profile_multicolor_rows():
    """
    Разбить профили с цветом вида "шампань, золото, коньяк" на отдельные строки.
    Возвращает (created, deleted).
    """
    created = 0
    deleted = 0
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, series, name, color, supplier, price_per_meter, photo_number, link, handle
                   FROM facades_profile ORDER BY id"""
            )
            rows = cur.fetchall()
            for row in rows:
                color = str(row.get('color') or '').strip()
                tokens = _split_color_tokens(color)
                if len(tokens) <= 1:
                    continue
                for token in tokens:
                    cur.execute(
                        """SELECT id FROM facades_profile
                           WHERE lower(series)=lower(%s) AND lower(name)=lower(%s) AND lower(color)=lower(%s)
                           LIMIT 1""",
                        (str(row.get('series') or '').strip(), str(row.get('name') or '').strip(), token)
                    )
                    if cur.fetchone():
                        continue
                    cur.execute(
                        """INSERT INTO facades_profile (series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)""",
                        (
                            str(row.get('series') or '').strip(),
                            str(row.get('name') or '').strip(),
                            token,
                            str(row.get('supplier') or '').strip(),
                            row.get('price_per_meter'),
                            str(row.get('photo_number') or '').strip(),
                            str(row.get('link') or '').strip(),
                            str(row.get('handle') or '').strip(),
                        )
                    )
                    created += 1
                cur.execute("DELETE FROM facades_profile WHERE id = %s", (int(row.get('id')),))
                deleted += 1
    return created, deleted


def _extract_continental_rows_from_xls(file_path):
    """
    Вернуть строки вида [{'article','color','price'}] из прайса Continental (.xls).
    Ожидается структура с 2 блоками колонок: (1,2,3) и (5,6,7).
    """
    try:
        import xlrd
    except ImportError:
        return [], "Установите xlrd: pip install xlrd"
    if not os.path.isfile(file_path):
        return [], "Файл не найден: %s" % file_path
    try:
        wb = xlrd.open_workbook(file_path)
        sh = wb.sheet_by_index(0)
    except Exception as e:
        return [], str(e)

    out = []
    # Два набора колонок на листе
    blocks = [(1, 2, 3), (5, 6, 7)]  # article, color, price
    last_article = {0: '', 1: ''}

    for r in range(sh.nrows):
        for bi, (c_article, c_color, c_price) in enumerate(blocks):
            article = sh.cell_value(r, c_article) if c_article < sh.ncols else ''
            color = sh.cell_value(r, c_color) if c_color < sh.ncols else ''
            price = sh.cell_value(r, c_price) if c_price < sh.ncols else ''

            article = str(article or '').strip()
            color = str(color or '').strip()
            # Цена может быть числом, строкой, "под заказ" и т.п.
            price_val = None
            if isinstance(price, (int, float)):
                price_val = round(float(price), 2)
            else:
                ptxt = str(price or '').strip().replace(' ', '').replace(',', '.')
                if ptxt and ptxt.lower() not in ('подзаказ', 'под', 'заказ'):
                    try:
                        price_val = round(float(ptxt), 2)
                    except ValueError:
                        price_val = None

            # Новая строка артикула
            if article:
                last_article[bi] = article
            article_use = last_article[bi]
            if not article_use or price_val is None:
                continue
            # Отсекаем не-артикульные строки
            if 'f' not in article_use.lower() and 'kr' not in article_use.lower():
                continue
            out.append({
                'article': article_use,
                'color': color,
                'price': price_val,
            })
    return out, None


def facades_import_continental_prices_from_xls(file_path, progress_cb=None):
    """
    Импорт цен Continental из xls:
    - ищет совпадения артикула по всем фасадным таблицам;
    - обновляет цены, а для профилей при наличии базового артикула добавляет недостающие цветовые варианты.
    Возвращает dict-отчёт.
    """
    rows, err = _extract_continental_rows_from_xls(file_path)
    if err:
        return {'error': err, 'updated': 0, 'created': 0, 'matches': [], 'not_found': []}

    profiles = facades_get_all_profiles()
    hinges = facades_get_all_hinges()
    screws = facades_get_all_screws()
    plates = facades_get_all_plates()
    seals = facades_get_all_seal()
    corners = facades_get_all_corners()

    datasets = [
        ('profiles', profiles, lambda i, p: facades_update_profile_price(i, p)),
        ('hinges', hinges, lambda i, p: facades_update_hinge_price(i, p)),
        ('screws', screws, lambda i, p: facades_update_screw_price(i, p)),
        ('plates', plates, lambda i, p: facades_update_plate_price(i, p)),
        ('seal', seals, lambda i, p: facades_update_seal_price(i, p)),
        ('corners', corners, lambda i, p: facades_update_corner_price(i, p)),
    ]

    updated = 0
    created = 0
    matches = []
    not_found = []

    total_rows = len(rows)
    for idx, src in enumerate(rows, start=1):
        if callable(progress_cb):
            try:
                progress_cb(idx, total_rows, src)
            except Exception:
                pass
        article = str(src.get('article') or '').strip()
        color = str(src.get('color') or '').strip()
        price = src.get('price')
        if price is None or not article:
            continue
        try:
            price_rub = _facade_price_rub_ceil(price)
        except (TypeError, ValueError):
            continue
        if price_rub is None:
            continue
        norm_article = _norm_article(article)
        found_any = False

        for kind, rows_db, updater in datasets:
            matched_rows = []
            for row in rows_db:
                hay = _norm_article("%s %s %s" % (row.get('number') or '', row.get('series') or '', row.get('name') or ''))
                if norm_article and norm_article in hay:
                    matched_rows.append(row)

            if not matched_rows:
                continue
            found_any = True

            # Для профилей: обновление существующего цвета + создание недостающего цветового варианта
            if kind == 'profiles':
                color_tokens = _split_color_tokens(color)
                if not color_tokens:
                    for row in matched_rows:
                        updater(int(row.get('id')), price_rub)
                        updated += 1
                        matches.append("%s: %s / %s -> %s" % (kind, article, '-', price_rub))
                    continue
                for token in color_tokens:
                    exact_color_rows = []
                    for row in matched_rows:
                        row_color_tokens = [x.lower() for x in _split_color_tokens(row.get('color'))]
                        if token.lower() in row_color_tokens or str(row.get('color') or '').strip().lower() == token.lower():
                            exact_color_rows.append(row)
                    if exact_color_rows:
                        for row in exact_color_rows:
                            updater(int(row.get('id')), price_rub)
                            updated += 1
                            matches.append("%s: %s / %s -> %s" % (kind, article, token, price_rub))
                    else:
                        base = matched_rows[0]
                        new_id = facades_add_profile_variant_from_base(int(base.get('id')), token, price_rub)
                        if new_id:
                            created += 1
                            matches.append("%s: создан цветовой вариант %s / %s -> %s" % (kind, article, token, price_rub))
                continue

            # Остальные таблицы: обновляем все совпавшие
            for row in matched_rows:
                updater(int(row.get('id')), price_rub)
                updated += 1
                matches.append("%s: %s / %s -> %s" % (kind, article, color or '-', price_rub))

        if not found_any:
            not_found.append("%s / %s / %s" % (article, color or '-', price_rub))

    split_created, split_deleted = facades_expand_profile_multicolor_rows()
    created += split_created

    return {
        'error': None,
        'updated': updated,
        'created': created,
        'matched_total': len(matches),
        'matches': matches,
        'not_found': not_found,
        'source_rows': len(rows),
        'split_deleted': split_deleted,
    }


def facades_get_profile_by_id(profile_id):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at FROM facades_profile WHERE id = %s",
                (int(profile_id),)
            )
            return _facades_apply_ceil_price_row(cur.fetchone(), "price_per_meter")


def facades_get_profiles_by_ids(profile_ids):
    """Один запрос: id профиля → строка facades_profile (для склада без N+1)."""
    ids = []
    for x in profile_ids or []:
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            continue
    ids = sorted(set(ids))
    if not ids:
        return {}
    placeholders = ",".join(["%s"] * len(ids))
    sql = (
        "SELECT id, series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at "
        "FROM facades_profile WHERE id IN (" + placeholders + ")"
    )
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(ids))
            rows = _facades_apply_ceil_price_rows(cur.fetchall() or [], "price_per_meter")
    return {int(r["id"]): r for r in rows if r.get("id") is not None}


def _facades_hinge_primary_catalog_sql():
    """
    Условие «основной каталог» (HARMONY PLUS): флаг в БД или вхождение в серию/название
    (подстрока, не только с начала строки — как в Excel).
    """
    return (
        "(catalog_primary IS TRUE OR "
        "COALESCE(TRIM(series), '') ILIKE '%%HARMONY PLUS%%' OR "
        "COALESCE(TRIM(name), '') ILIKE '%%HARMONY PLUS%%')"
    )


def facades_get_all_hinges(catalog_filter=None):
    """
    Список петель.
    catalog_filter:
      None — все (вкладка «Цены», админ);
      'primary' — основной каталог (HARMONY PLUS);
      'others' — всё остальное (не HARMONY PLUS и без catalog_primary).
    """
    where = ""
    prim = _facades_hinge_primary_catalog_sql()
    if catalog_filter == "primary":
        where = " WHERE " + prim
    elif catalog_filter == "others":
        where = " WHERE NOT (" + prim + ")"
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, number, series, name, color, supplier, price, price_updated_at, item, photo_number, link, catalog_primary
                   FROM facades_hinges"""
                + where
                + " ORDER BY series, name"
            )
            return _facades_apply_ceil_price_rows(cur.fetchall(), "price")


def facades_update_hinge_price(hinge_id, price):
    p = _facade_price_rub_ceil(price)
    if p is None:
        return
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE facades_hinges SET price = %s, price_updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                (p, int(hinge_id))
            )


def facades_get_hinge_by_id(hinge_id):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, number, series, name, color, supplier, price, price_updated_at, item, photo_number, link, catalog_primary
                   FROM facades_hinges WHERE id = %s""",
                (int(hinge_id),)
            )
            return _facades_apply_ceil_price_row(cur.fetchone(), "price")


def _facades_get_all_items(table_name):
    """Общий запрос для facades_screws, facades_plates, facades_seal, facades_corners."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT id, number, series, name, color, supplier, price, price_updated_at, photo_number, link
                   FROM %s ORDER BY series, name""" % table_name
            )
            return _facades_apply_ceil_price_rows(cur.fetchall(), "price")


def _facades_update_item_price(table_name, item_id, price):
    p = _facade_price_rub_ceil(price)
    if p is None:
        return
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE %s SET price = %%s, price_updated_at = CURRENT_TIMESTAMP WHERE id = %%s" % table_name,
                (p, int(item_id))
            )


def facades_get_all_screws():
    return _facades_get_all_items('facades_screws')


def facades_update_screw_price(screw_id, price):
    _facades_update_item_price('facades_screws', screw_id, price)


def facades_get_all_plates():
    return _facades_get_all_items('facades_plates')


def facades_update_plate_price(plate_id, price):
    _facades_update_item_price('facades_plates', plate_id, price)


def facades_get_all_seal():
    return _facades_get_all_items('facades_seal')


def facades_update_seal_price(seal_id, price):
    _facades_update_item_price('facades_seal', seal_id, price)


def facades_get_all_corners():
    return _facades_get_all_items('facades_corners')


def facades_update_corner_price(corner_id, price):
    _facades_update_item_price('facades_corners', corner_id, price)


def facades_get_all_angle_seal():
    """Сводная таблица 'уголки_уплотнитель' для фасадов (админские цены)."""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, item_type, variant, unit, price, price_updated_at, link, created_at
                FROM facades_angle_seal
                ORDER BY item_type, variant
                """
            )
            return _facades_apply_ceil_price_rows(cur.fetchall(), "price")


def facades_update_angle_seal_price(angle_seal_id, price):
    p = _facade_price_rub_ceil(price)
    if p is None:
        return
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE facades_angle_seal
                SET price = %s, price_updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (p, int(angle_seal_id)),
            )


def facades_insert_hinge(name, link=None, price=None):
    p = _facade_price_rub_ceil(price) if price is not None else None
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO facades_hinges (name, link, price, price_updated_at, catalog_primary)
                   VALUES (%s, %s, %s, CASE WHEN %s IS NOT NULL THEN CURRENT_TIMESTAMP ELSE NULL END, FALSE) RETURNING id""",
                (str(name or '').strip(), (link or '').strip() or None, p, p)
            )
            row = cur.fetchone()
            return row['id'] if row else None


def facades_import_harmony_plus_hinges_from_seed(delete_existing_primary=True, fetch_prices=True):
    """
    Разовая загрузка HARMONY PLUS в facades_hinges: при fetch_prices — цена с сайта МДМ, иначе из списка.
    Строки помечаются catalog_primary=TRUE. Существующие петли не удаляются.
    Повторный запуск с delete_existing_primary=True сначала удаляет строки основного каталога.
    """
    if _mp_dir not in sys.path:
        sys.path.insert(0, _mp_dir)
    from FASAD.harmony_plus_hinges_data import iter_harmony_plus_hinge_rows

    rows = list(iter_harmony_plus_hinge_rows())
    if not rows:
        return 0, "Нет строк для импорта"

    fetch_price_from_mdm_url = None
    if fetch_prices:
        try:
            from FASAD.mdm_parser import fetch_price_from_mdm_url as _fetch
            fetch_price_from_mdm_url = _fetch
        except ImportError:
            fetch_price_from_mdm_url = None

    with get_connection() as conn:
        with conn.cursor() as cur:
            if delete_existing_primary:
                cur.execute(
                    "DELETE FROM facades_hinges WHERE " + _facades_hinge_primary_catalog_sql()
                )
            for r in rows:
                url = (r.get("link") or "").strip()
                price = None
                if fetch_price_from_mdm_url and url:
                    price = fetch_price_from_mdm_url(url)
                if price is None:
                    price = r.get("sheet_price")
                try:
                    price = _facade_price_rub_ceil(price) if price is not None else None
                except (TypeError, ValueError):
                    price = None
                cur.execute(
                    """INSERT INTO facades_hinges
                       (number, series, name, color, supplier, price, price_updated_at, item, photo_number, link, catalog_primary)
                       VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, '', %s, %s, TRUE)""",
                    (
                        str(r.get("number") or "").strip(),
                        "HARMONY PLUS",
                        str(r.get("name") or "").strip(),
                        str(r.get("color") or "").strip(),
                        str(r.get("supplier") or "").strip(),
                        price,
                        str(r.get("photo_number") or "").strip(),
                        url or None,
                    ),
                )
    return len(rows), None


def facades_ensure_harmony_plus_hinges_if_empty():
    """
    Если в БД нет ни одной петли основного каталога (HARMONY PLUS), подставляет сида из кода.
    Цены из таблицы (без HTTP при старте). Имеет смысл вызывать после ensure_facades_tables и импорта Excel.
    """
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM facades_hinges WHERE "
                + _facades_hinge_primary_catalog_sql()
            )
            n = (cur.fetchone() or {}).get("cnt") or 0
    if n > 0:
        return
    facades_import_harmony_plus_hinges_from_seed(
        delete_existing_primary=False,
        fetch_prices=False,
    )


def get_default_profiles_excel_path():
    """Профили: Profil_new.xlsx в корне MIRROR_CUT, иначе FASAD/Профиль.xlsx."""
    new_p = os.path.join(_root, "Profil_new.xlsx")
    if os.path.isfile(new_p):
        return new_p
    legacy = os.path.join(_mp_dir, "FASAD", "Профиль.xlsx")
    return legacy if os.path.isfile(legacy) else new_p


def get_default_facades_bundle_excel_path():
    """Петли/винты/прочее: классический FASAD/Профиль.xlsx (если есть)."""
    p = os.path.join(_mp_dir, "FASAD", "Профиль.xlsx")
    return p if os.path.isfile(p) else ""


def facades_import_hinges_from_excel(file_path, replace=False):
    """
    Импорт петель из Excel (лист «Петли»).
    Колонки: серия, название, цвет, поставщик, стоимость/цена, номер фото, ссылка (артикула в файле нет).
    replace=True: сначала очистить таблицу facades_hinges, затем вставить.
    Возвращает (count_inserted, error_message or None).
    """
    try:
        import openpyxl
    except ImportError:
        return 0, "Установите openpyxl: pip install openpyxl"
    if not os.path.isfile(file_path):
        return 0, "Файл не найден: %s" % file_path
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = None
        for s in wb.sheetnames:
            if s and 'петл' in s.lower():
                ws = wb[s]
                break
        if ws is None:
            wb.close()
            return 0, "Лист «Петли» не найден в файле"
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0, values_only=True))
        wb.close()
    except Exception as e:
        return 0, str(e)
    if not rows:
        return 0, "Лист «Петли» пустой"
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    header_joined = ' '.join(header).lower()
    col = {}
    for i, h in enumerate(header):
        h_lower = (h or '').lower().strip()
        if not h_lower:
            continue
        if 'сери' in h_lower:
            col['series'] = i
        # Имя может быть «Название», «Наименование» и т.п.
        if 'назван' in h_lower or 'наимен' in h_lower:
            col['name'] = i
        if 'цвет' in h_lower:
            col['color'] = i
        if 'поставщ' in h_lower:
            col['supplier'] = i
        if 'стоим' in h_lower or 'цена' in h_lower or 'цен' in h_lower or 'cost' in h_lower:
            col['price'] = i
        if 'фото' in h_lower or ('номер' in h_lower and 'фото' in header_joined):
            col['photo'] = i
        if 'ссылка' in h_lower or h_lower == 'link':
            col['link'] = i
        if (h or '').strip() == '№' or (h_lower == '№') or ('номер' in h_lower and 'фото' not in h_lower) or h_lower == 'no':
            col['number'] = i
    if 'name' not in col:
        return 0, "Не найдена колонка «Название» в листе Петли"
    count = 0
    with get_connection() as conn:
        with conn.cursor() as cur:
            if replace:
                cur.execute("DELETE FROM facades_hinges")
            for row in rows[1:]:
                if not row:
                    continue
                name_val = (row[col['name']] if col['name'] < len(row) else None) or ''
                if not str(name_val).strip():
                    continue
                number_val = row[col['number']] if col.get('number') is not None and col['number'] < len(row) else ''
                number_val = str(number_val or '').strip() if number_val is not None else ''
                # Серия по номеру: 1–18 → NIMBUS, 19–36 → PRISMA\FUTURIS (как в списках из Excel)
                try:
                    num_int = int(number_val) if number_val else 0
                    if 1 <= num_int <= 18:
                        series_val = 'NIMBUS'
                    elif 19 <= num_int <= 36:
                        series_val = 'PRISMA\\FUTURIS'
                    else:
                        series_val = row[col['series']] if col.get('series') is not None and col['series'] < len(row) else ''
                except (ValueError, TypeError):
                    series_val = row[col['series']] if col.get('series') is not None and col['series'] < len(row) else ''
                if not str(series_val or '').strip():
                    series_val = row[col['series']] if col.get('series') is not None and col['series'] < len(row) else ''
                color_val = row[col['color']] if col.get('color') is not None and col['color'] < len(row) else ''
                supplier_val = row[col['supplier']] if col.get('supplier') is not None and col['supplier'] < len(row) else ''
                link_val = row[col['link']] if col.get('link') is not None and col['link'] < len(row) else ''
                price_val = None
                if col.get('price') is not None and col['price'] < len(row):
                    raw = row[col['price']]
                    if raw is not None and str(raw).strip():
                        try:
                            s = str(raw).strip().replace(',', '.')
                            price_val = _facade_price_rub_ceil(float(s))
                        except (TypeError, ValueError):
                            pass
                photo_val = row[col['photo']] if col.get('photo') is not None and col['photo'] < len(row) else ''
                cur.execute(
                    """INSERT INTO facades_hinges (number, series, name, color, supplier, price, price_updated_at, item, photo_number, link, catalog_primary)
                       VALUES (%s, %s, %s, %s, %s, %s, CASE WHEN %s IS NOT NULL THEN CURRENT_TIMESTAMP ELSE NULL END, '', %s, %s, FALSE)""",
                    (number_val, str(series_val or '').strip() if series_val is not None else '', str(name_val).strip(), str(color_val or '').strip(),
                     str(supplier_val or '').strip(), price_val, price_val,
                     str(photo_val or '').strip(), str(link_val or '').strip() or None)
                )
                count += 1
    return count, None


def _facades_import_items_from_excel(file_path, sheet_predicate, table_name, empty_sheet_error, replace=False):
    """
    Общий импорт фасадных элементов (винты, площадки, уплотнитель, уголки) из Excel.
    Структура колонок аналогична листу «Петли».
    """
    try:
        import openpyxl
    except ImportError:
        return 0, "Установите openpyxl: pip install openpyxl"
    if not os.path.isfile(file_path):
        return 0, "Файл не найден: %s" % file_path
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = None
        for s in wb.sheetnames:
            if s and sheet_predicate(s.lower()):
                ws = wb[s]
                break
        if ws is None:
            wb.close()
            return 0, empty_sheet_error
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0, values_only=True))
        wb.close()
    except Exception as e:
        return 0, str(e)
    if not rows:
        return 0, empty_sheet_error
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    header_joined = ' '.join(header).lower()
    col = {}
    for i, h in enumerate(header):
        h_lower = (h or '').lower().strip()
        if not h_lower:
            continue
        if 'сери' in h_lower:
            col['series'] = i
        # Имя может быть «Название», «Наименование», «Описание» и т.п.
        if 'назван' in h_lower or 'наимен' in h_lower or 'описан' in h_lower:
            col['name'] = i
        if 'цвет' in h_lower:
            col['color'] = i
        if 'поставщ' in h_lower:
            col['supplier'] = i
        if 'стоим' in h_lower or 'цена' in h_lower or 'цен' in h_lower or 'cost' in h_lower:
            col['price'] = i
        if 'фото' in h_lower or ('номер' in h_lower and 'фото' in header_joined):
            col['photo'] = i
        if 'ссылка' in h_lower or h_lower == 'link':
            col['link'] = i
        if (h or '').strip() == '№' or (h_lower == '№') or ('номер' in h_lower and 'фото' not in h_lower) or h_lower == 'no':
            col['number'] = i

    # Если колонку имени автоматически не нашли, постараемся выбрать её эвристикой
    if 'name' not in col:
        excluded = {'series', 'color', 'supplier', 'price', 'photo', 'link', 'number'}
        used_indices = set(v for k, v in col.items() if k in excluded)
        name_idx = None
        for i, h in enumerate(header):
            if i in used_indices:
                continue
            h_lower = (h or '').lower().strip()
            if not h_lower:
                continue
            # Игнорируем заведомо служебные колонки
            if any(x in h_lower for x in ('высота', 'мм', 'кол-во', 'количество', 'шт/')):
                continue
            name_idx = i
            break
        if name_idx is None:
            # В крайнем случае используем ту же колонку, что и series/number, либо первую
            if 'series' in col:
                name_idx = col['series']
            elif 'number' in col:
                name_idx = col['number']
            else:
                name_idx = 0
        col['name'] = name_idx
    count = 0
    with get_connection() as conn:
        with conn.cursor() as cur:
            if replace:
                cur.execute("DELETE FROM " + table_name)
            for row in rows[1:]:
                if not row:
                    continue
                name_val = (row[col['name']] if col['name'] < len(row) else None) or ''
                if not str(name_val).strip():
                    continue
                number_val = row[col['number']] if col.get('number') is not None and col['number'] < len(row) else ''
                number_val = str(number_val or '').strip() if number_val is not None else ''
                series_val = row[col['series']] if col.get('series') is not None and col['series'] < len(row) else ''
                color_val = row[col['color']] if col.get('color') is not None and col['color'] < len(row) else ''
                supplier_val = row[col['supplier']] if col.get('supplier') is not None and col['supplier'] < len(row) else ''
                link_val = row[col['link']] if col.get('link') is not None and col['link'] < len(row) else ''
                price_val = None
                if col.get('price') is not None and col['price'] < len(row):
                    raw = row[col['price']]
                    if raw is not None and str(raw).strip():
                        try:
                            s = str(raw).strip().replace(',', '.')
                            price_val = _facade_price_rub_ceil(float(s))
                        except (TypeError, ValueError):
                            pass
                photo_val = row[col['photo']] if col.get('photo') is not None and col['photo'] < len(row) else ''
                cur.execute(
                    "INSERT INTO " + table_name + " (number, series, name, color, supplier, price, price_updated_at, photo_number, link) "
                    "VALUES (%s, %s, %s, %s, %s, %s, CASE WHEN %s IS NOT NULL THEN CURRENT_TIMESTAMP ELSE NULL END, %s, %s)",
                    (
                        number_val,
                        str(series_val or '').strip() if series_val is not None else '',
                        str(name_val).strip(),
                        str(color_val or '').strip(),
                        str(supplier_val or '').strip(),
                        price_val,
                        price_val,
                        str(photo_val or '').strip(),
                        str(link_val or '').strip() or None,
                    ),
                )
                count += 1
    return count, None


def facades_import_screws_from_excel(file_path, replace=False):
    return _facades_import_items_from_excel(
        file_path,
        lambda s: 'винт' in s or 'шуруп' in s,
        'facades_screws',
        "Лист «Винты» не найден или пустой",
        replace,
    )


def facades_import_plates_from_excel(file_path, replace=False):
    return _facades_import_items_from_excel(
        file_path,
        lambda s: 'площад' in s,
        'facades_plates',
        "Лист «Площадки» не найден или пустой",
        replace,
    )


def facades_import_seal_from_excel(file_path, replace=False):
    return _facades_import_items_from_excel(
        file_path,
        lambda s: 'уплотнит' in s or 'уплотн' in s,
        'facades_seal',
        "Лист «Уплотнитель» не найден или пустой",
        replace,
    )


def facades_import_corners_from_excel(file_path, replace=False):
    return _facades_import_items_from_excel(
        file_path,
        lambda s: 'уголк' in s or 'угол' in s,
        'facades_corners',
        "Лист «Уголки» не найден или пустой",
        replace,
    )


def facades_ensure_import_from_excel_once():
    """
    Один раз подтягивает данные фасадов из Excel:
    если таблица пуста, то делает импорт; если уже есть записи, Excel больше не трогаем.
    """
    profile_path = get_default_profiles_excel_path()
    bundle_path = get_default_facades_bundle_excel_path()
    has_prof = profile_path and os.path.isfile(profile_path)
    has_bundle = bool(bundle_path) and os.path.isfile(bundle_path)
    if not has_prof and not has_bundle:
        return
    hinge_xl = bundle_path if has_bundle else (profile_path if has_prof else "")

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    (SELECT COUNT(*) FROM facades_profile) AS pc,
                    (SELECT COUNT(*) FROM facades_hinges) AS hc,
                    (SELECT COUNT(*) FROM facades_screws) AS sc,
                    (SELECT COUNT(*) FROM facades_plates) AS plc,
                    (SELECT COUNT(*) FROM facades_seal) AS sec,
                    (SELECT COUNT(*) FROM facades_corners) AS cc
                """
            )
            r0 = cur.fetchone() or {}
            profiles_empty = (r0.get("pc") or 0) == 0
            hinges_empty = (r0.get("hc") or 0) == 0
            screws_empty = (r0.get("sc") or 0) == 0
            plates_empty = (r0.get("plc") or 0) == 0
            seal_empty = (r0.get("sec") or 0) == 0
            corners_empty = (r0.get("cc") or 0) == 0

    if profiles_empty and has_prof:
        facades_import_profiles_from_excel(profile_path, replace=True)
    if hinges_empty and hinge_xl:
        facades_import_hinges_from_excel(hinge_xl, replace=True)
    if screws_empty and hinge_xl:
        facades_import_screws_from_excel(hinge_xl, replace=True)
    if plates_empty and hinge_xl:
        facades_import_plates_from_excel(hinge_xl, replace=True)
    if seal_empty and hinge_xl:
        facades_import_seal_from_excel(hinge_xl, replace=True)
    if corners_empty and hinge_xl:
        facades_import_corners_from_excel(hinge_xl, replace=True)


def facades_import_profiles_from_excel(file_path, replace=False):
    """
    Импорт профилей из Excel (лист «Профиль», структура Profil_new.xlsx).
    replace=True: очистить facades_profile и вставить заново.
    Возвращает (count_inserted, error_message or None).
    """
    rows_flat, err = _iter_profil_excel_profile_rows(file_path)
    if err:
        return 0, err
    if not rows_flat:
        return 0, "Нет строк для импорта профилей"
    count = 0
    with get_connection() as conn:
        with conn.cursor() as cur:
            if replace:
                cur.execute("DELETE FROM facades_profile")
            for r in rows_flat:
                cur.execute(
                    """INSERT INTO facades_profile (series, name, color, supplier, price_per_meter, photo_number, link, handle, price_updated_at)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)""",
                    (
                        r["series"],
                        r["name"],
                        r["color"],
                        r["supplier"],
                        r["price_per_meter"],
                        r["photo_number"],
                        r["link"] or "",
                        (r.get("handle") or "0")[:255],
                    ),
                )
                count += 1
    return count, None


def facades_update_profile_prices_from_excel(file_path):
    """
    Обновление только цен (и price_updated_at).
    Поддерживается:
      • простой шаблон — первая строка: Серия, Название, Цвет, Цена/м;
      • формат Profil_new (лист «Профиль», колонки A–H).
    Совпадение по серия + название + цвет (без учёта регистра, trim).
    Возвращает dict: error, updated, not_found, source_rows.
    """
    rows_simple, err_simple = _iter_simple_profile_price_rows(file_path)
    if rows_simple:
        return _facades_apply_profile_price_updates(rows_simple)
    rows_old, err_old = _iter_profil_excel_profile_rows(file_path)
    if err_old:
        err_msg = err_simple or err_old
        return {"error": err_msg, "updated": 0, "not_found": [], "source_rows": 0}
    slim = []
    for r in rows_old:
        slim.append(
            {
                "series": r["series"],
                "name": r["name"],
                "color": r["color"],
                "price_per_meter": r["price_per_meter"],
            }
        )
    return _facades_apply_profile_price_updates(slim)
